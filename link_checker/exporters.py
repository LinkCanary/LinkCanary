"""Export formats for LinkCanary reports.

Supports: CSV, JSON, MDX, Excel, PDF, Google Sheets
"""

import csv
import json
import logging
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Optional

import pandas as pd

logger = logging.getLogger(__name__)


class ExportFormat:
    """Available export formats."""
    CSV = "csv"
    JSON = "json"
    MDX = "mdx"
    EXCEL = "xlsx"
    PDF = "pdf"
    GOOGLE_SHEETS = "sheets"


class ReportExporter:
    """Export reports to multiple formats."""

    def __init__(self, df: pd.DataFrame, summary: dict = None):
        """
        Initialize exporter with report data.
        
        Args:
            df: DataFrame containing the report data
            summary: Optional summary statistics
        """
        self.df = df
        self.summary = summary or {}
        self.timestamp = datetime.now().isoformat()

    def export(self, output_path: str, format: str = None) -> str:
        """
        Export report to specified format.
        
        Args:
            output_path: Output file path (extension used if format not specified)
            format: Export format (csv, json, mdx, xlsx, pdf, sheets)
        
        Returns:
            Path to exported file
        """
        if format is None:
            format = Path(output_path).suffix.lstrip('.').lower()
        
        exporters = {
            ExportFormat.CSV: self.export_csv,
            ExportFormat.JSON: self.export_json,
            ExportFormat.MDX: self.export_mdx,
            ExportFormat.EXCEL: self.export_excel,
            ExportFormat.PDF: self.export_pdf,
            ExportFormat.GOOGLE_SHEETS: self.export_google_sheets,
        }
        
        exporter = exporters.get(format)
        if not exporter:
            raise ValueError(f"Unsupported export format: {format}. "
                           f"Supported: {list(exporters.keys())}")
        
        return exporter(output_path)

    def export_csv(self, output_path: str) -> str:
        """Export to CSV format."""
        self.df.to_csv(output_path, index=False)
        logger.info(f"CSV report saved to {output_path}")
        return output_path

    def export_json(self, output_path: str) -> str:
        """Export to JSON format with metadata."""
        data = {
            "metadata": {
                "exported_at": self.timestamp,
                "total_issues": len(self.df),
                "summary": self.summary,
            },
            "issues": self.df.to_dict(orient="records"),
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"JSON report saved to {output_path}")
        return output_path

    def export_mdx(self, output_path: str) -> str:
        """Export to MDX (Markdown + JSX) for documentation sites.
        
        Generates a formatted markdown file with React components
        suitable for Next.js, Docusaurus, or similar documentation frameworks.
        """
        lines = [
            "---",
            'title: "LinkCanary Report"',
            f'date: "{self.timestamp}"',
            f'totalIssues: {len(self.df)}',
            "---",
            "",
            "## Summary",
            "",
            "| Priority | Count |",
            "|----------|-------|",
            f"| Critical | {self.summary.get('critical', 0)} |",
            f"| High | {self.summary.get('high', 0)} |",
            f"| Medium | {self.summary.get('medium', 0)} |",
            f"| Low | {self.summary.get('low', 0)} |",
            "",
            "## Issues",
            "",
        ]
        
        if self.df.empty:
            lines.append("*No issues found.*")
        else:
            # Group by priority
            for priority in ['critical', 'high', 'medium', 'low']:
                priority_df = self.df[self.df['priority'] == priority]
                if priority_df.empty:
                    continue
                
                lines.append(f"### {priority.title()} Priority ({len(priority_df)})")
                lines.append("")
                
                for _, row in priority_df.head(20).iterrows():
                    lines.append(f"#### [{row['link_url'][:80]}]({row['link_url']})")
                    lines.append("")
                    lines.append(f"- **Status**: {row['status_code']}")
                    lines.append(f"- **Type**: {row['issue_type']}")
                    lines.append(f"- **Occurrences**: {row['occurrence_count']}")
                    
                    if row.get('source_page') and row['source_page'] != 'multiple':
                        lines.append(f"- **Source**: [{row['source_page'][:50]}]({row['source_page']})")
                    
                    if row.get('recommended_fix'):
                        lines.append(f"- **Fix**: {row['recommended_fix']}")
                    
                    lines.append("")
                
                if len(priority_df) > 20:
                    lines.append(f"*...and {len(priority_df) - 20} more {priority} priority issues*")
                    lines.append("")
        
        # Add interactive component placeholder for MDX
        lines.extend([
            "## Export Options",
            "",
            "```jsx",
            "<LinkCanaryExport data={issues} />",
            "```",
        ])
        
        content = "\n".join(lines)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"MDX report saved to {output_path}")
        return output_path

    def export_excel(self, output_path: str) -> str:
        """Export to Excel format with formatting.
        
        Requires: openpyxl or xlsxwriter
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            # Fallback to xlsxwriter
            return self._export_excel_xlsxwriter(output_path)
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ["LinkCanary Report", ""],
            ["Generated", self.timestamp],
            ["Total Issues", len(self.df)],
            ["", ""],
            ["Priority", "Count"],
            ["Critical", self.summary.get('critical', 0)],
            ["High", self.summary.get('high', 0)],
            ["Medium", self.summary.get('medium', 0)],
            ["Low", self.summary.get('low', 0)],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Style summary
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A5'].font = Font(bold=True)
        ws_summary['B5'].font = Font(bold=True)
        
        # Issues sheet
        ws_issues = wb.create_sheet("Issues")
        
        # Add headers
        headers = list(self.df.columns)
        ws_issues.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, header in enumerate(headers, 1):
            cell = ws_issues.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row in dataframe_to_rows(self.df, index=False, header=False):
            ws_issues.append(row)
        
        # Style priority cells
        priority_colors = {
            'critical': 'FF0000',
            'high': 'FF6600',
            'medium': 'FFCC00',
            'low': '0066CC',
        }
        
        priority_col = headers.index('priority') + 1 if 'priority' in headers else None
        if priority_col:
            for row_num in range(2, len(self.df) + 2):
                priority_val = ws_issues.cell(row=row_num, column=priority_col).value
                if priority_val in priority_colors:
                    fill = PatternFill(
                        start_color=priority_colors[priority_val],
                        end_color=priority_colors[priority_val],
                        fill_type="solid"
                    )
                    ws_issues.cell(row=row_num, column=priority_col).fill = fill
        
        # Auto-fit columns
        for col in ws_issues.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_issues.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        wb.save(output_path)
        logger.info(f"Excel report saved to {output_path}")
        return output_path

    def _export_excel_xlsxwriter(self, output_path: str) -> str:
        """Fallback Excel export using xlsxwriter."""
        try:
            import xlsxwriter
        except ImportError:
            raise ImportError(
                "Excel export requires 'openpyxl' or 'xlsxwriter'. "
                "Install with: pip install openpyxl xlsxwriter"
            )
        
        workbook = xlsxwriter.Workbook(output_path)
        
        # Summary sheet
        ws_summary = workbook.add_worksheet("Summary")
        
        title_format = workbook.add_format({'bold': True, 'font_size': 14})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': 'white'})
        
        ws_summary.write('A1', 'LinkCanary Report', title_format)
        ws_summary.write('A2', 'Generated', header_format)
        ws_summary.write('B2', self.timestamp)
        ws_summary.write('A3', 'Total Issues', header_format)
        ws_summary.write('B3', len(self.df))
        
        # Issues sheet
        ws_issues = workbook.add_worksheet("Issues")
        
        # Write headers
        for col, header in enumerate(self.df.columns):
            ws_issues.write(0, col, header, header_format)
        
        # Write data
        for row_idx, row in enumerate(self.df.itertuples(index=False), start=1):
            for col_idx, value in enumerate(row):
                ws_issues.write(row_idx, col_idx, str(value) if value else '')
        
        workbook.close()
        logger.info(f"Excel report saved to {output_path}")
        return output_path

    def export_pdf(self, output_path: str) -> str:
        """Export to PDF format.
        
        Requires: reportlab or weasyprint
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            return self._export_pdf_weasyprint(output_path)
        
        doc = SimpleDocTemplate(output_path, pagesize=letter,
                               rightMargin=0.5*inch, leftMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=12,
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph("LinkCanary Report", title_style))
        elements.append(Paragraph(f"Generated: {self.timestamp}", styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Summary table
        summary_data = [
            ['Priority', 'Count'],
            ['Critical', str(self.summary.get('critical', 0))],
            ['High', str(self.summary.get('high', 0))],
            ['Medium', str(self.summary.get('medium', 0))],
            ['Low', str(self.summary.get('low', 0))],
            ['Total', str(len(self.df))],
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 24))
        
        # Issues table (top 50 to avoid huge PDFs)
        if not self.df.empty:
            elements.append(Paragraph("Top Issues", styles['Heading2']))
            elements.append(Spacer(1, 6))
            
            # Prepare data for table
            headers = ['URL', 'Status', 'Priority', 'Type']
            data = [headers]
            
            for _, row in self.df.head(50).iterrows():
                data.append([
                    row['link_url'][:60] + ('...' if len(row['link_url']) > 60 else ''),
                    str(row['status_code']),
                    row['priority'],
                    row['issue_type'],
                ])
            
            issues_table = Table(data, colWidths=[4*inch, 0.8*inch, 1*inch, 1.5*inch])
            
            # Color coding for priorities
            priority_colors = {
                'critical': colors.red,
                'high': colors.orange,
                'medium': colors.yellow,
                'low': colors.lightblue,
            }
            
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]
            
            # Add priority colors
            for i, priority in enumerate(self.df.head(50)['priority']):
                if priority in priority_colors:
                    table_style.append(('BACKGROUND', (2, i+1), (2, i+1), priority_colors[priority]))
            
            issues_table.setStyle(TableStyle(table_style))
            elements.append(issues_table)
            
            if len(self.df) > 50:
                elements.append(Spacer(1, 12))
                elements.append(Paragraph(
                    f"...and {len(self.df) - 50} more issues (see CSV export for full list)",
                    styles['Italic']
                ))
        
        doc.build(elements)
        logger.info(f"PDF report saved to {output_path}")
        return output_path

    def _export_pdf_weasyprint(self, output_path: str) -> str:
        """Fallback PDF export using weasyprint from HTML."""
        try:
            from weasyprint import HTML
        except ImportError:
            raise ImportError(
                "PDF export requires 'reportlab' or 'weasyprint'. "
                "Install with: pip install reportlab weasyprint"
            )
        
        # Generate HTML content
        html_content = self._generate_pdf_html()
        
        # Convert to PDF
        HTML(string=html_content).write_pdf(output_path)
        logger.info(f"PDF report saved to {output_path}")
        return output_path

    def _generate_pdf_html(self) -> str:
        """Generate HTML content for PDF export."""
        rows_html = ""
        for _, row in self.df.head(100).iterrows():
            priority_class = f"priority-{row['priority']}"
            rows_html += f"""
            <tr class="{priority_class}">
                <td>{row['link_url'][:80]}</td>
                <td>{row['status_code']}</td>
                <td>{row['priority']}</td>
                <td>{row['issue_type']}</td>
            </tr>
            """
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                h1 {{ color: #333; }}
                .summary {{ margin-bottom: 20px; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; font-size: 12px; }}
                th {{ background-color: #4472C4; color: white; }}
                .priority-critical {{ background-color: #ffcccc; }}
                .priority-high {{ background-color: #ffe6cc; }}
                .priority-medium {{ background-color: #ffffcc; }}
                .priority-low {{ background-color: #cce6ff; }}
            </style>
        </head>
        <body>
            <h1>LinkCanary Report</h1>
            <p>Generated: {self.timestamp}</p>
            
            <div class="summary">
                <h2>Summary</h2>
                <p>Critical: {self.summary.get('critical', 0)} | 
                   High: {self.summary.get('high', 0)} | 
                   Medium: {self.summary.get('medium', 0)} | 
                   Low: {self.summary.get('low', 0)}</p>
            </div>
            
            <table>
                <tr>
                    <th>URL</th>
                    <th>Status</th>
                    <th>Priority</th>
                    <th>Type</th>
                </tr>
                {rows_html}
            </table>
            
            {f'<p>...and {len(self.df) - 100} more issues</p>' if len(self.df) > 100 else ''}
        </body>
        </html>
        """

    def export_google_sheets(self, output_path: str = None, 
                            spreadsheet_id: str = None,
                            credentials_path: str = None,
                            sheet_name: str = "LinkCanary Report") -> str:
        """
        Export to Google Sheets.
        
        Requires: google-auth, google-auth-oauthlib, google-auth-httplib2, google-api-python-client
        
        Args:
            output_path: Not used for Google Sheets (kept for API consistency)
            spreadsheet_id: Existing spreadsheet ID (creates new if not provided)
            credentials_path: Path to OAuth credentials JSON file
            sheet_name: Name for the sheet tab
        
        Returns:
            URL to the Google Sheets document
        
        Note: This requires OAuth setup. See documentation for setup instructions.
        """
        try:
            from google.oauth2.service_account import Credentials
            from googleapiclient.discovery import build
        except ImportError:
            raise ImportError(
                "Google Sheets export requires google-api-python-client. "
                "Install with: pip install google-api-python-client google-auth"
            )
        
        if not credentials_path:
            raise ValueError(
                "Google Sheets export requires credentials_path. "
                "Set GOOGLE_APPLICATION_CREDENTIALS env var or pass credentials_path."
            )
        
        # Authenticate
        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        creds = Credentials.from_service_account_file(credentials_path, scopes=scopes)
        service = build('sheets', 'v4', credentials=creds)
        
        if spreadsheet_id:
            # Append to existing spreadsheet
            spreadsheet_id = spreadsheet_id
        else:
            # Create new spreadsheet
            spreadsheet = {
                'properties': {'title': f'LinkCanary Report {datetime.now().strftime("%Y-%m-%d")}'}
            }
            result = service.spreadsheets().create(body=spreadsheet).execute()
            spreadsheet_id = result['spreadsheetId']
        
        # Prepare data
        headers = list(self.df.columns)
        data = [headers] + self.df.values.tolist()
        
        # Write data
        body = {
            'values': data
        }
        
        result = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f'{sheet_name}!A1',
            valueInputOption='RAW',
            body=body
        ).execute()
        
        url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        logger.info(f"Google Sheets updated: {url}")
        return url

    def get_available_formats(self) -> list[str]:
        """Return list of available export formats."""
        return [ExportFormat.CSV, ExportFormat.JSON, ExportFormat.MDX, 
                ExportFormat.EXCEL, ExportFormat.PDF]


def detect_format(path: str) -> str:
    """Detect export format from file extension."""
    ext = Path(path).suffix.lower()
    format_map = {
        '.csv': ExportFormat.CSV,
        '.json': ExportFormat.JSON,
        '.mdx': ExportFormat.MDX,
        '.md': ExportFormat.MDX,
        '.xlsx': ExportFormat.EXCEL,
        '.xls': ExportFormat.EXCEL,
        '.pdf': ExportFormat.PDF,
    }
    return format_map.get(ext, ExportFormat.CSV)
